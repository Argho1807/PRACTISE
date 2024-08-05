""" BOX PACKING """

""" 

PROBLEM STATEMENT -
A big box of length L and height H 
Small boxes of length l and height h
These can be placed in l-h or h-l direction inside the big box
Maximise number of small boxes that can be placed

"""

import gurobipy as gp
from gurobipy import GRB
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class Optimizer():
    def __init__(self):
        self.readInput()
        self.model = gp.Model('BoxPacking')
        self.variables()
        self.constraints()
        self.objectiveFunction()
        self.setRunParams()
        self.printOutputs()

    def readInput(self):
        self.LENGTH = 60
        self.HEIGHT = 35
        self.length = 11
        self.height = 3

    def variables(self):

        self.ifBoxKeptLengthwise = self.model.addVars(
            [x for x in range(1, self.HEIGHT-self.height+2)]
            , [y for y in range(1, self.LENGTH-self.length+2)]
            , vtype=GRB.BINARY
            , name=f"if_box_kept_lengthwise")

        self.ifBoxKeptHeightwise = self.model.addVars(
            [x for x in range(1, self.HEIGHT-self.length+2)]
            , [y for y in range(1, self.LENGTH-self.height+2)]
            , vtype=GRB.BINARY
            , name=f"if_box_kept_heightwise")

    def constraints(self):
        """ CONSTRAINTS """
        """ Only maximum one small box starting inside 1*1 grid of big box """

        for x in range(1, self.HEIGHT+1):
            for y in range(1, self.LENGTH+1):
                self.model.addConstr(

                    gp.quicksum(self.ifBoxKeptLengthwise[i, j]
                                for i in range(max(1, x-self.height+1), x+1)
                                for j in range(max(1, y-self.length+1), y+1)
                                if (i, j) in self.ifBoxKeptLengthwise)

                    + gp.quicksum(self.ifBoxKeptHeightwise[i, j]
                                  for i in range(max(1, x-self.length+1), x+1)
                                  for j in range(max(1, y-self.height+1), y+1)
                                  if (i, j) in self.ifBoxKeptHeightwise)

                    <= 1, f"max_one_small_box_start_grid_x_{x}_y_{y}")

    def objectiveFunction(self):
        self.numSmallBoxes = self.ifBoxKeptLengthwise.sum('*', '*')+self.ifBoxKeptHeightwise.sum('*', '*')

    def setRunParams(self):

        self.model.setObjective(self.numSmallBoxes, GRB.MAXIMIZE)
        self.model.setParam('TimeLimit', 300)
        # self.model.setParam('NoRelHeurTime', 200)
        self.model.setParam('MIPGap', 0)
        self.model.write('outputs/lp_file.lp')
        self.model.optimize()

    def borders(self):
        self.border = Border(
            left=Side(border_style='thick', color='000000'),
            right=Side(border_style='thick', color='000000'),
            top=Side(border_style='thick', color='000000'),
            bottom=Side(border_style='thick', color='000000')
        )

        self.left_border = Border(
            left=Side(border_style='thick', color='000000')
        )

        self.right_border = Border(
            right=Side(border_style='thick', color='000000')
        )

        self.top_border = Border(
            top=Side(border_style='thick', color='000000')
        )

        self.bottom_border = Border(
            bottom=Side(border_style='thick', color='000000')
        )

        self.left_top_border = Border(
            left=Side(border_style='thick', color='000000'),
            top=Side(border_style='thick', color='000000')
        )

        self.left_bottom_border = Border(
            left=Side(border_style='thick', color='000000'),
            bottom=Side(border_style='thick', color='000000')
        )

        self.right_bottom_border = Border(
            right=Side(border_style='thick', color='000000'),
            bottom=Side(border_style='thick', color='000000')
        )

        self.right_top_border = Border(
            right=Side(border_style='thick', color='000000'),
            top=Side(border_style='thick', color='000000')
        )

    def fills(self):
        self.yellowFill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.orangeFill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    def applyOuterBorders(self):
        self.ws.cell(row=1, column=1).border = self.left_top_border
        self.ws.cell(row=self.HEIGHT, column=1).border = self.left_bottom_border
        self.ws.cell(row=self.HEIGHT, column=self.LENGTH).border = self.right_bottom_border
        self.ws.cell(row=1, column=self.LENGTH).border = self.right_top_border

        for row in range(2, self.HEIGHT):
            self.ws.cell(row=row, column=1).border = self.left_border
            self.ws.cell(row=row, column=self.LENGTH).border = self.right_border

        for column in range(2, self.LENGTH):
            self.ws.cell(row=1, column=column).border = self.top_border
            self.ws.cell(row=self.HEIGHT, column=column).border = self.bottom_border

    def fixRowColSize(self):
        rowSize = 20
        colSize = rowSize * 1.33 / 7

        for row in range(1, self.HEIGHT + 1):
            self.ws.row_dimensions[row].height = rowSize

        for col in range(1, self.LENGTH + 1):
            # col_letter = chr(64 + col)
            col_letter = get_column_letter(col)
            self.ws.column_dimensions[col_letter].width = colSize

    def fillOutputValues(self):
        smallBoxCount = 0

        for i, j in self.ifBoxKeptLengthwise:
            if self.ifBoxKeptLengthwise[i, j].X > 0.1:
                smallBoxCount += 1
                self.ws.merge_cells(start_row=i, start_column=j, end_row=i + self.height - 1,
                                    end_column=j + self.length - 1)
                cell = self.ws.cell(row=i, column=j)
                cell.value = smallBoxCount
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

                for row in self.ws[i: i + self.height - 1]:
                    for cell in row[j - 1: j + self.length - 1]:
                        cell.fill = self.yellowFill
                        cell.border = self.border

        for i, j in self.ifBoxKeptHeightwise:
            if self.ifBoxKeptHeightwise[i, j].X > 0.1:
                smallBoxCount += 1
                self.ws.merge_cells(start_row=i, start_column=j, end_row=i + self.length - 1,
                                    end_column=j + self.height - 1)
                cell = self.ws.cell(row=i, column=j)
                cell.value = smallBoxCount
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

                for row in self.ws[i: i + self.length - 1]:
                    for cell in row[j - 1: j + self.height - 1]:
                        cell.fill = self.orangeFill
                        cell.border = self.border

    def printOutputs(self):
        if self.model.STATUS not in [GRB.INFEASIBLE, GRB.UNBOUNDED, GRB.INF_OR_UNBD]:
            print("Optimal Solution") if self.model.STATUS == GRB.OPTIMAL else print("Feasible Solution")
            print(f"Objective - {self.model.objVal} small boxes")
            print(f"{self.ifBoxKeptLengthwise.sum('*', '*').getValue()} boxes kept in l-h orientation")
            print(f"{self.ifBoxKeptHeightwise.sum('*', '*').getValue()} boxes kept in h-l orientation")
            for v in self.model.getVars():
                if v.X > 0.5:
                    print(f'{v.VarName} - {v.X}')

            self.wb = Workbook()
            self.ws = self.wb.active

            self.borders()
            self.fills()
            self.fixRowColSize()
            self.applyOuterBorders()
            self.fillOutputValues()

            self.wb.save(f"outputs/Box Packing_{self.LENGTH}_{self.HEIGHT}_{self.length}_{self.height}.xlsx")

Optimizer()